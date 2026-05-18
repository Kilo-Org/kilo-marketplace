#!/usr/bin/env python3
"""
edit_xlsx.py — Edit an existing .xlsx / .xlsm file using openpyxl.

Usage:
  python edit_xlsx.py INPUT OUTPUT [options]

Options:
  --replace OLD NEW               Replace all cell text values containing OLD with NEW
  --set-cell SHEET ROW COL VAL    Set a cell to a plain value (row/col are 1-based integers)
  --set-formula SHEET ROW COL F   Set a cell to a formula (must start with =)
  --rename-sheet OLD NEW          Rename a sheet

Examples:
  python edit_xlsx.py in.xlsx out.xlsx --replace "Draft" "Final"
  python edit_xlsx.py in.xlsx out.xlsx --set-cell Sheet1 1 3 "Q1 Revenue"
  python edit_xlsx.py in.xlsx out.xlsx --set-formula Sheet1 10 3 "=SUM(C2:C9)"
  python edit_xlsx.py in.xlsx out.xlsx --rename-sheet "Sheet1" "Summary"
  python edit_xlsx.py in.xlsx out.xlsx --replace "TBD" "100" --set-formula Summary 5 2 "=SUM(B2:B4)"
"""
from __future__ import annotations

import argparse
import sys

try:
    import openpyxl
except ImportError:
    print('Error: openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(1)


def _get_sheet(
    wb: openpyxl.Workbook, name: str
) -> openpyxl.worksheet.worksheet.Worksheet:
    if name not in wb.sheetnames:
        raise ValueError(
            f'Sheet {name!r} not found. Available: {wb.sheetnames}'
        )
    return wb[name]


def replace(wb: openpyxl.Workbook, old: str, new: str) -> int:
    """Replace all cell string values containing OLD with NEW. Returns count."""
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and old in cell.value:
                    cell.value = cell.value.replace(old, new)
                    count += 1
    return count


def set_cell(
    wb: openpyxl.Workbook, sheet: str, row: int, col: int, value: str
) -> None:
    """Set a cell to a plain value, coercing to int or float when possible."""
    ws = _get_sheet(wb, sheet)
    coerced: str | int | float = value
    try:
        coerced = int(value)
    except ValueError:
        try:
            coerced = float(value)
        except ValueError:
            pass
    ws.cell(row=row, column=col, value=coerced)


def set_formula(
    wb: openpyxl.Workbook, sheet: str, row: int, col: int, formula: str
) -> None:
    """Set a cell to a formula string (must start with =)."""
    if not formula.startswith('='):
        raise ValueError(f'Formula must start with =, got: {formula!r}')
    ws = _get_sheet(wb, sheet)
    ws.cell(row=row, column=col, value=formula)


def rename_sheet(wb: openpyxl.Workbook, old: str, new: str) -> None:
    """Rename a sheet."""
    ws = _get_sheet(wb, old)
    ws.title = new


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument('input',  help='Source .xlsx')
    ap.add_argument('output', help='Destination .xlsx')
    ap.add_argument(
        '--replace',
        nargs=2, metavar=('OLD', 'NEW'),
        action='append', default=[],
    )
    ap.add_argument(
        '--set-cell',
        nargs=4, metavar=('SHEET', 'ROW', 'COL', 'VALUE'),
        action='append', default=[], dest='set_cell',
    )
    ap.add_argument(
        '--set-formula',
        nargs=4, metavar=('SHEET', 'ROW', 'COL', 'FORMULA'),
        action='append', default=[], dest='set_formula',
    )
    ap.add_argument(
        '--rename-sheet',
        nargs=2, metavar=('OLD', 'NEW'),
        action='append', default=[], dest='rename_sheet',
    )
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input)

    for old, new in args.replace:
        n = replace(wb, old, new)
        print(f'Replaced {n} occurrence(s) of {old!r} with {new!r}')

    for sheet, row, col, value in args.set_cell:
        set_cell(wb, sheet, int(row), int(col), value)
        print(f'Set {sheet}!R{row}C{col} = {value!r}')

    for sheet, row, col, formula in args.set_formula:
        set_formula(wb, sheet, int(row), int(col), formula)
        print(f'Set {sheet}!R{row}C{col} = {formula}')

    for old, new in args.rename_sheet:
        rename_sheet(wb, old, new)
        print(f'Renamed sheet {old!r} -> {new!r}')

    wb.save(args.output)
    print(f'Saved: {args.output}')


if __name__ == '__main__':
    main()
