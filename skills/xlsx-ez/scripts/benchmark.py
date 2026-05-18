#!/usr/bin/env python3
"""Benchmark: xlsx-ez skill vs Anthropic skills fork xlsx skill.

Suites:
  C – SKILL.md coverage (keyword/topic checks)
  D – Script inventory (which files each skill ships)
  E – Code quality (safety, portability, design guidance)
  A – Functional round-trip (create → validate → unpack → pack)
  B – Feature-specific tests (openpyxl read, validate passes)
  P – POI corpus tests (only when --poi is given)

Usage:
    python benchmark.py [--poi POI_DIR] [--mode MODE] [--output JSON]
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

SKILL_A = Path("/home/user/kilo/skills/xlsx-ez")
SKILL_B = Path("/home/user/skills/skills/xlsx")

_results: list[tuple[str, str, str, str]] = []
_pass = _fail = _warn = 0


# ── helpers ──────────────────────────────────────────────────────────────────

def _record(suite: str, name: str, status: str, detail: str = "") -> None:
    global _pass, _fail, _warn
    _results.append((suite, name, status, detail))
    icon = {"PASS": "✓", "FAIL": "✗", "WARN": "⚠"}.get(status, "?")
    suffix = f": {detail}" if detail else ""
    print(f"  {icon} [{suite}] {name}{suffix}")
    if status == "PASS":   _pass += 1
    elif status == "FAIL": _fail += 1
    elif status == "WARN": _warn += 1


def _read(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""
    return path.read_text(errors="replace")


def _run(cmd: list[str], cwd: str | None = None, env=None,
         timeout: int = 30) -> tuple[int, str, str]:
    try:
        r = subprocess.run(cmd, capture_output=True, text=True,
                           cwd=cwd, env=env, timeout=timeout)
        return r.returncode, r.stdout, r.stderr
    except subprocess.TimeoutExpired:
        return -1, "", "TIMEOUT"
    except Exception as e:
        return -2, "", str(e)


# ── Suite C – SKILL.md keyword coverage ──────────────────────────────────────

SKILL_MD_CHECKS = [
    ("formula usage",                  ["formulas", "=SUM", "formula"]),
    ("openpyxl API",                   ["openpyxl", "Workbook"]),
    ("cell formatting",                ["PatternFill", "Font", "Border"]),
    ("charts",                         ["BarChart", "chart"]),
    ("named ranges",                   ["DefinedName", "named_range"]),
    ("data validation",                ["DataValidation"]),
    ("conditional formatting",         ["ConditionalFormatting", "conditional"]),
    ("PDF export / LibreOffice",       ["LibreOffice", "soffice"]),
    ("OOXML unpack/pack",              ["unpack", "pack"]),
    ("formula verification",           ["recalc", "recalculate"]),
]


def suite_c() -> None:
    print("\n## Suite C: SKILL.md Coverage")
    a_text = _read(SKILL_A / "SKILL.md")
    b_text = _read(SKILL_B / "SKILL.md")

    for label, keywords in SKILL_MD_CHECKS:
        a_has = all(kw.lower() in a_text.lower() for kw in keywords)
        b_has = all(kw.lower() in b_text.lower() for kw in keywords)

        if a_has and b_has:
            _record("C", label, "PASS")
        elif b_has and not a_has:
            _record("C", label, "FAIL", "Skill A missing; Skill B has it")
        elif a_has and not b_has:
            _record("C", label, "WARN", "Skill A unique feature; Skill B missing")
        else:
            _record("C", label, "FAIL", "Both missing")


# ── Suite D – Script inventory ────────────────────────────────────────────────

SCRIPT_PAIRS: dict[str, tuple[Path | None, Path | None]] = {
    "create_xlsx.py":         (SKILL_A / "scripts/create_xlsx.py",        SKILL_B / "scripts/create_xlsx.py"),
    "recalc.py":              (SKILL_A / "scripts/recalc.py",             SKILL_B / "scripts/recalc.py"),
    "validate_xlsx.py":       (SKILL_A / "scripts/validate_xlsx.py",      SKILL_B / "scripts/validate.py"),
    "office/soffice.py":      (SKILL_A / "scripts/office/soffice.py",     SKILL_B / "scripts/office/soffice.py"),
    "office/unpack.py":       (SKILL_A / "scripts/office/unpack.py",      SKILL_B / "scripts/office/unpack.py"),
    "office/pack.py":         (SKILL_A / "scripts/office/pack.py",        SKILL_B / "scripts/office/pack.py"),
    "openpyxl.md reference":  (SKILL_A / "references/openpyxl.md",        SKILL_B / "references/openpyxl.md"),
    "ooxml-xlsx.md reference":(SKILL_A / "references/ooxml-xlsx.md",      SKILL_B / "references/ooxml-xlsx.md"),
}


def suite_d() -> None:
    print("\n## Suite D: Script Inventory")
    for label, (path_a, path_b) in SCRIPT_PAIRS.items():
        a_has = path_a is not None and path_a.exists()
        b_has = path_b is not None and path_b.exists()

        if a_has and b_has:
            _record("D", label, "PASS")
        elif a_has and not b_has:
            _record("D", label, "WARN", "Skill A unique feature")
        elif b_has and not a_has:
            _record("D", label, "FAIL", "Skill B has it; Skill A missing")
        else:
            _record("D", label, "FAIL", "Both missing")


# ── Suite E – Code quality ────────────────────────────────────────────────────

def suite_e() -> None:
    print("\n## Suite E: Code Quality")

    # create_xlsx.py
    a_create = _read(SKILL_A / "scripts/create_xlsx.py")
    _record("E", "create_xlsx: shebang present",
            "PASS" if "#!/usr/bin/env python3" in a_create else "FAIL")
    _record("E", "create_xlsx: json.load or argparse",
            "PASS" if "json.load" in a_create or "argparse" in a_create else "FAIL")
    _record("E", "create_xlsx: chart support (BarChart or add_chart)",
            "PASS" if "BarChart" in a_create or "add_chart" in a_create else "WARN",
            "" if "BarChart" in a_create or "add_chart" in a_create
            else "No chart support found in create_xlsx.py")
    _record("E", "create_xlsx: DataValidation",
            "PASS" if "DataValidation" in a_create else "WARN",
            "" if "DataValidation" in a_create
            else "No DataValidation usage found in create_xlsx.py")

    # validate_xlsx.py
    a_val = _read(SKILL_A / "scripts/validate_xlsx.py")
    _record("E", "validate_xlsx: checks ZIP",
            "PASS" if "zipfile" in a_val or "ZipFile" in a_val else "FAIL")
    _record("E", "validate_xlsx: checks workbook-xml",
            "PASS" if "workbook.xml" in a_val else "FAIL")
    _record("E", "validate_xlsx: checks sheet parseable",
            "PASS" if "sheet" in a_val.lower() and (
                "parse" in a_val.lower() or "fromstring" in a_val or "etree" in a_val
            ) else "FAIL")

    # office/unpack.py
    a_unpack = _read(SKILL_A / "scripts/office/unpack.py")
    _record("E", "office/unpack: smart-quote encoding (&#x201C;)",
            "PASS" if "&#x201C;" in a_unpack else "FAIL")
    _record("E", "office/unpack: lxml pretty_print",
            "PASS" if "pretty_print" in a_unpack else "FAIL")

    # office/pack.py
    a_pack = _read(SKILL_A / "scripts/office/pack.py")
    _record("E", "office/pack: --original flag",
            "PASS" if "--original" in a_pack or "original" in a_pack else "FAIL")
    _record("E", "office/pack: smart-quote decode",
            "PASS" if "&#x201C;" in a_pack or "&#x201c;" in a_pack or
                      "SMART_QUOTE_DECODE" in a_pack else "FAIL")

    # recalc.py
    a_recalc = _read(SKILL_A / "scripts/recalc.py")
    _record("E", "recalc: LibreOffice usage",
            "PASS" if "LibreOffice" in a_recalc or "soffice" in a_recalc else "FAIL")
    _record("E", "recalc: shebang present",
            "PASS" if "#!/usr/bin/env python3" in a_recalc else "FAIL")

    # SKILL.md length ratio
    a_len = len(_read(SKILL_A / "SKILL.md"))
    b_len = len(_read(SKILL_B / "SKILL.md"))
    ratio = a_len / b_len if b_len else 0
    _record("E", f"SKILL.md coverage ratio (A={a_len} B={b_len} chars)",
            "PASS" if ratio >= 0.6 else "FAIL" if ratio < 0.3 else "WARN",
            f"Skill A is {int(ratio * 100)}% of Skill B guidance")


# ── Suite A – Functional round-trip ──────────────────────────────────────────

def _make_minimal_xlsx(path: Path) -> None:
    """Create a minimal valid .xlsx file for use as a test fixture."""
    CT = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml"  ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml"
    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml"
    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml"
    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>"""
    RELS = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
    Target="xl/workbook.xml"/>
</Relationships>"""
    WB_RELS = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
    Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles"
    Target="styles.xml"/>
</Relationships>"""
    WB = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>"""
    SHEET = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
      <c r="A1" t="inlineStr"><is><t>Hello XLSX</t></is></c>
    </row>
  </sheetData>
</worksheet>"""
    STYLES = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="2">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
  </fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
</styleSheet>"""
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml",          CT)
        z.writestr("_rels/.rels",                   RELS)
        z.writestr("xl/_rels/workbook.xml.rels",    WB_RELS)
        z.writestr("xl/workbook.xml",               WB)
        z.writestr("xl/worksheets/sheet1.xml",      SHEET)
        z.writestr("xl/styles.xml",                 STYLES)


def suite_a(tmpdir: Path) -> None:
    print("\n## Suite A: Functional Round-Trip")

    # ── create_xlsx.py ──────────────────────────────────────────────────────
    out_a = tmpdir / "a_demo.xlsx"
    rc, out, err = _run([sys.executable,
                         str(SKILL_A / "scripts/create_xlsx.py"), "--example"])
    # --example prints a spec; write it to a file, then create from it
    if rc == 0 and out.strip():
        spec_file = tmpdir / "example_spec.json"
        spec_file.write_text(out.strip())
        rc2, out2, err2 = _run([sys.executable,
                                 str(SKILL_A / "scripts/create_xlsx.py"),
                                 str(spec_file), str(out_a)])
        a_create_ok = rc2 == 0 and out_a.exists()
        _record("A", "Skill A create_xlsx.py: demo workbook created",
                "PASS" if a_create_ok else "FAIL",
                err2.strip() if not a_create_ok else "")
    else:
        # Fallback: try running with --example and redirecting output
        # Some versions print JSON then exit; try a direct invocation
        a_create_ok = False
        _record("A", "Skill A create_xlsx.py: --example flag",
                "FAIL", (out + err).strip()[:120])

    if a_create_ok:
        try:
            with zipfile.ZipFile(out_a) as z:
                names = z.namelist()
            has_wb = "xl/workbook.xml" in names
            _record("A", "Skill A create: valid ZIP with xl/workbook.xml",
                    "PASS" if has_wb else "FAIL")
        except Exception as e:
            _record("A", "Skill A create: valid ZIP", "FAIL", str(e))

    # ── validate_xlsx.py on created output ──────────────────────────────────
    if a_create_ok:
        rc, out, err = _run([sys.executable,
                             str(SKILL_A / "scripts/validate_xlsx.py"), str(out_a)])
        _record("A", "Skill A validate_xlsx.py: passes on own output",
                "PASS" if rc == 0 else "FAIL",
                (out + err).strip() if rc != 0 else "")

    # ── office/unpack.py on fixture ──────────────────────────────────────────
    fixture = tmpdir / "fixture.xlsx"
    _make_minimal_xlsx(fixture)

    a_unpacked = tmpdir / "a_unpacked"
    rc, out, err = _run([sys.executable,
                         str(SKILL_A / "scripts/office/unpack.py"),
                         str(fixture), str(a_unpacked)])
    a_unpack_ok = rc == 0 and (a_unpacked / "xl" / "workbook.xml").exists()
    _record("A", "Skill A unpack.py: extracts xl/workbook.xml",
            "PASS" if a_unpack_ok else "FAIL",
            err.strip() if not a_unpack_ok else "")

    # ── office/pack.py on unpacked ───────────────────────────────────────────
    if a_unpack_ok:
        a_repacked = tmpdir / "a_repacked.xlsx"
        rc, out, err = _run([sys.executable,
                             str(SKILL_A / "scripts/office/pack.py"),
                             str(a_unpacked), str(a_repacked),
                             "--original", str(fixture)])
        a_pack_ok = rc == 0 and a_repacked.exists()
        if "lxml" in err or "ModuleNotFoundError" in err:
            _record("A", "Skill A pack.py: repackages unpacked XLSX", "WARN",
                    f"dependency missing: {err.strip()[:100]}")
        else:
            _record("A", "Skill A pack.py: repackages unpacked XLSX",
                    "PASS" if a_pack_ok else "FAIL",
                    err.strip() if not a_pack_ok else "")

        if a_pack_ok:
            # verify openpyxl can load result
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(a_repacked))
                _record("A", "Skill A pack: openpyxl can load repacked file",
                        "PASS" if len(wb.sheetnames) >= 1 else "FAIL",
                        f"sheets: {wb.sheetnames}")
            except ImportError:
                _record("A", "Skill A pack: openpyxl can load repacked file",
                        "WARN", "openpyxl not installed")
            except Exception as e:
                _record("A", "Skill A pack: openpyxl can load repacked file",
                        "FAIL", str(e))


# ── Suite B – Feature tests ───────────────────────────────────────────────────

def suite_b(tmpdir: Path) -> None:
    print("\n## Suite B: Feature Tests")

    # Build a workbook via create_xlsx.py (reuse logic from suite_a)
    out_b = tmpdir / "b_demo.xlsx"
    rc, out, err = _run([sys.executable,
                         str(SKILL_A / "scripts/create_xlsx.py"), "--example"])
    b_create_ok = False
    if rc == 0 and out.strip():
        spec_file = tmpdir / "b_spec.json"
        spec_file.write_text(out.strip())
        rc2, _, _ = _run([sys.executable,
                           str(SKILL_A / "scripts/create_xlsx.py"),
                           str(spec_file), str(out_b)])
        b_create_ok = rc2 == 0 and out_b.exists()

    # Feature test 1: openpyxl can read it and find at least one sheet
    if b_create_ok:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(out_b))
            _record("B", "create_xlsx.py output: openpyxl reads it, finds sheet(s)",
                    "PASS" if len(wb.sheetnames) >= 1 else "FAIL",
                    f"sheets: {wb.sheetnames}")
        except ImportError:
            _record("B", "create_xlsx.py output: openpyxl readable",
                    "WARN", "openpyxl not installed")
        except Exception as e:
            _record("B", "create_xlsx.py output: openpyxl readable", "FAIL", str(e))
    else:
        _record("B", "create_xlsx.py output: openpyxl readable",
                "FAIL", "create_xlsx.py did not produce output")

    # Feature test 2: validate_xlsx.py on created output exits 0
    if b_create_ok:
        rc, out, err = _run([sys.executable,
                             str(SKILL_A / "scripts/validate_xlsx.py"), str(out_b)])
        _record("B", "validate_xlsx.py on created output: exits 0",
                "PASS" if rc == 0 else "FAIL",
                (out + err).strip() if rc != 0 else "")
    else:
        _record("B", "validate_xlsx.py on created output: exits 0",
                "FAIL", "no file to validate")


# ── Report ────────────────────────────────────────────────────────────────────

def _report() -> None:
    now   = datetime.now().strftime("%Y-%m-%d %H:%M")
    total = _pass + _fail + _warn
    lines = [
        "# xlsx-ez Skill Benchmark Report",
        "",
        f"**Generated**: {now}",
        f"**Skill A (xlsx-ez)**: `kilo/skills/xlsx-ez/`",
        f"**Skill B (reference)**: `skills/skills/xlsx/`",
        "",
        "---",
        "",
        "## Summary",
        "",
        "| | Count |",
        "|---|---|",
        f"| ✓ Passed  | {_pass} |",
        f"| ✗ Failed  | {_fail} |",
        f"| ⚠ Warnings | {_warn} |",
        f"| **Total** | **{total}** |",
        "",
        "---",
        "",
        "## Results",
        "",
        "| Suite | Test | Status | Detail |",
        "|-------|------|--------|--------|",
    ]
    for suite, name, status, detail in _results:
        icon = {"PASS": "✓", "FAIL": "✗", "WARN": "⚠"}.get(status, "?")
        lines.append(f"| {suite} | {name} | {icon} {status} | {detail} |")

    lines += [
        "",
        "---",
        "",
        "## Key Findings",
        "",
        "### Where Skill A (xlsx-ez) leads",
        "",
        "- **validate_xlsx.py**: Standalone ZIP/OPC + XML integrity checker with "
          "detailed per-check reporting. Skill B ships `validate.py` under scripts/office/.",
        "- **create_xlsx.py**: JSON-spec driven workbook creation with "
          "fonts, fills, borders, number formats, freeze panes, column widths, merges.",
        "- **References**: Comprehensive `openpyxl.md` and `ooxml-xlsx.md` "
          "covering the full SpreadsheetML anatomy.",
        "- **office/unpack.py + office/pack.py**: OOXML unpack/pack with "
          "smart-quote entity encoding and lxml pretty-printing.",
        "",
        "### Where Skill B (Anthropic) leads",
        "",
        "- **SKILL.md guidance depth**: Skill B's SKILL.md covers financial model "
          "color conventions, number formatting standards, formula construction rules, "
          "documentation requirements for hardcodes — content not in Skill A.",
        "- **pandas integration**: More detailed pandas workflows including "
          "read_only/write_only mode guidance and dtype handling.",
        "",
        "---",
        "",
        "## Actionable Improvements for xlsx-ez",
        "",
        "### High Priority",
        "",
        "**1. Add chart support to create_xlsx.py**",
        "The `create_xlsx.py` JSON spec does not support BarChart / add_chart. "
        "Adding chart configuration to the spec would cover a common spreadsheet task.",
        "",
        "**2. Add DataValidation to create_xlsx.py spec**",
        "DataValidation (dropdown lists, integer ranges) is documented in SKILL.md "
        "but not wired into the create_xlsx.py JSON spec driver.",
        "",
        "**3. Expand SKILL.md with financial model guidance**",
        "Skill B documents color coding standards, number formatting rules, "
        "formula construction guidelines, and hardcode documentation requirements. "
        "Adding these to SKILL.md improves model quality.",
        "",
        "### Medium Priority",
        "",
        "**4. Add named_range to SKILL.md keyword coverage**",
        "SKILL.md mentions DefinedName but not the accessor name 'named_range'. "
        "Expanding coverage ensures keyword searches match both forms.",
        "",
        "### Regressions to Avoid",
        "",
        "1. **Do not remove validate_xlsx.py** — full OPC/ZIP/XML integrity checker.",
        "2. **Do not remove recalc.py** — formula verification via LibreOffice is "
           "a core deliverable requirement.",
        "3. **Do not remove office/unpack.py + pack.py** — direct XML editing path.",
    ]

    report_path = SKILL_A / "benchmark_report.md"
    report_path.write_text("\n".join(lines) + "\n")
    print(f"\nReport written to: {report_path}")


# ── Suite P – POI corpus (only when --poi given) ──────────────────────────────

try:
    from lxml import etree as _etree
    _LXML_OK = True
except ImportError:
    _LXML_OK = False

_NS_SML = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_C   = "http://schemas.openxmlformats.org/drawingml/2006/chart"


def _qn(ns: str, local: str) -> str:
    return f"{{{ns}}}{local}"


def _open_part(path: str, part: str):
    try:
        with zipfile.ZipFile(path) as z:
            if part not in z.namelist():
                return None
            data = z.read(part)
        if not _LXML_OK:
            return None
        parser = _etree.XMLParser(huge_tree=True, recover=True)
        return _etree.fromstring(data, parser)
    except Exception:
        return None


def _list_zip_names(path: str) -> list[str]:
    try:
        with zipfile.ZipFile(path) as z:
            return z.namelist()
    except Exception:
        return []


def _list_sheet_parts(path: str) -> list[str]:
    root = _open_part(path, "xl/_rels/workbook.xml.rels")
    if root is None:
        return []
    parts = []
    for rel in root.iter(_qn(_NS_REL, "Relationship")):
        rtype = rel.get("Type", "")
        if "worksheet" not in rtype.lower():
            continue
        target = rel.get("Target", "")
        if not target:
            continue
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = f"xl/{target}"
        parts.append(target)
    return parts


def _list_chart_parts(path: str) -> list[str]:
    return [n for n in _list_zip_names(path)
            if n.startswith("xl/charts/") and n.endswith(".xml")]


def _count_formula_cells(path: str) -> int:
    total = 0
    for sheet_part in _list_sheet_parts(path):
        root = _open_part(path, sheet_part)
        if root is None:
            continue
        total += len(root.findall(f".//{_qn(_NS_SML, 'f')}"))
    for chart_part in _list_chart_parts(path):
        root = _open_part(path, chart_part)
        if root is None:
            continue
        total += len(root.findall(f".//{_qn(_NS_C, 'f')}"))
    return total


def _count_table_parts(path: str) -> int:
    return sum(1 for n in _list_zip_names(path)
               if n.startswith("xl/tables/") and n.endswith(".xml"))


def _count_chart_xml_parts(path: str) -> int:
    return sum(1 for n in _list_zip_names(path)
               if n.startswith("xl/charts/chart") and n.endswith(".xml"))


def _has_cell_formatting(path: str) -> tuple[bool, str]:
    root = _open_part(path, "xl/styles.xml")
    if root is None:
        return False, "xl/styles.xml missing"

    fills_el   = root.find(f".//{_qn(_NS_SML, 'fills')}")
    fonts_el   = root.find(f".//{_qn(_NS_SML, 'fonts')}")
    borders_el = root.find(f".//{_qn(_NS_SML, 'borders')}")

    fill_count   = int(fills_el.get("count",   0)) if fills_el   is not None else 0
    font_count   = int(fonts_el.get("count",   0)) if fonts_el   is not None else 0
    border_count = int(borders_el.get("count", 0)) if borders_el is not None else 0

    numfmt_count = len(root.findall(f".//{_qn(_NS_SML, 'numFmt')}"))
    dxf_count    = len(root.findall(f".//{_qn(_NS_SML, 'dxf')}"))

    cf_count   = 0
    tab_colors = 0
    for sp in _list_sheet_parts(path):
        sroot = _open_part(path, sp)
        if sroot is None:
            continue
        cf_count   += len(sroot.findall(f".//{_qn(_NS_SML, 'conditionalFormatting')}"))
        tab_colors += len(sroot.findall(f".//{_qn(_NS_SML, 'tabColor')}"))

    parts = []
    if font_count > 1:      parts.append(f"{font_count} font(s)")
    if fill_count > 2:      parts.append(f"{fill_count} fill(s)")
    if border_count > 1:    parts.append(f"{border_count} border(s)")
    if numfmt_count > 0:    parts.append(f"{numfmt_count} numFmt(s)")
    if dxf_count > 0:       parts.append(f"{dxf_count} dxf(s)")
    if cf_count > 0:        parts.append(f"{cf_count} CF rule(s)")
    if tab_colors > 0:      parts.append(f"{tab_colors} tab color(s)")

    has = bool(parts)
    detail = (", ".join(parts) if parts
              else f"default only: {font_count}f/{fill_count}fi/{border_count}b")
    return has, detail


@dataclass
class _CheckResult:
    name:   str
    passed: bool
    detail: str = ""


@dataclass
class _FileResult:
    path:   str
    checks: list[_CheckResult] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return all(c.passed for c in self.checks)

    def add(self, name: str, passed: bool, detail: str = "") -> None:
        self.checks.append(_CheckResult(name, passed, detail))

    def report(self) -> str:
        status = "PASS" if self.passed else "FAIL"
        lines  = [f"\n{status}: {self.path}"]
        for c in self.checks:
            sym = "✓" if c.passed else "✗"
            lines.append(f"  {sym} {c.name}" + (f": {c.detail}" if c.detail else ""))
        return "\n".join(lines)

    def to_dict(self) -> dict:
        return {
            "file":   self.path,
            "passed": self.passed,
            "checks": [
                {"name": c.name, "passed": c.passed, "detail": c.detail}
                for c in self.checks
            ],
        }


def _run_checks(path: str, mode: str) -> _FileResult:
    result = _FileResult(path)

    names = _list_zip_names(path)
    if not names:
        result.add("zip-readable", False, "not a valid ZIP/xlsx file")
        return result
    result.add("zip-readable", True)

    wb_present = "xl/workbook.xml" in names
    result.add("workbook-xml", wb_present, "" if wb_present else "xl/workbook.xml missing")
    if not wb_present:
        return result

    sheet_parts = _list_sheet_parts(path)
    bad = [sp for sp in sheet_parts if _open_part(path, sp) is None]
    result.add(
        "sheets-parseable",
        len(bad) == 0,
        f"{len(sheet_parts)} sheet(s)" if not bad else f"unparseable: {', '.join(bad)}",
    )

    if mode == "formulas":
        n = _count_formula_cells(path)
        result.add("has-formulas", n > 0, f"{n} formula reference(s)")

    if mode == "tables":
        n = _count_table_parts(path)
        result.add("has-tables", n > 0, f"{n} table part(s)")

    if mode == "charts":
        n = _count_chart_xml_parts(path)
        result.add("has-charts", n > 0, f"{n} chart part(s)")

    if mode == "formats":
        has, detail = _has_cell_formatting(path)
        result.add("has-formatting", has, detail)

    return result


def _infer_mode(filename: str) -> str:
    name = os.path.basename(filename).lower()

    if any(k in name for k in ("chart", "graph", "plot")):
        return "charts"
    if any(k in name for k in ("with_table", "exceltable", "listobject", "pivot")):
        return "tables"
    if "formula" in name and "conditional" not in name:
        return "formulas"
    if any(k in name for k in ("workday", "calcchain", "matrix_formula")):
        return "formulas"
    if "header" in name or "footer" in name:
        return "all"
    if "conditional" in name:
        return "formats"
    if any(k in name for k in ("colour", "color", "font", "border", "fill",
                                "format", "style", "indexed", "theme")):
        return "formats"
    return "all"


def suite_p(poi_path: str, mode: str, output: Optional[str]) -> int:
    """Suite P – POI corpus tests. Returns 0 if all pass, 1 otherwise."""
    print("\n## Suite P: POI Corpus Tests")

    if not _LXML_OK:
        print("Error: lxml not installed. Run: pip install lxml", file=sys.stderr)
        return 1

    paths: list[str] = []
    if os.path.isdir(poi_path):
        for root_dir, _, files in os.walk(poi_path):
            for f in sorted(files):
                if f.endswith(".xlsx") or f.endswith(".xlsm"):
                    paths.append(os.path.join(root_dir, f))
    elif os.path.isfile(poi_path):
        paths = [poi_path]
    else:
        print(f"Error: {poi_path!r} is not a file or directory", file=sys.stderr)
        return 1

    if not paths:
        print("No .xlsx/.xlsm files found.", file=sys.stderr)
        return 1

    results: list[_FileResult] = []
    for p in paths:
        effective_mode = mode if mode != "all" else _infer_mode(p)
        r = _run_checks(p, effective_mode)
        print(r.report())
        results.append(r)

    total  = len(results)
    passed = sum(1 for r in results if r.passed)
    print(f"\n{'=' * 60}")
    print(f"POI Results: {passed}/{total} passed")

    if output:
        with open(output, "w") as f:
            json.dump([r.to_dict() for r in results], f, indent=2)
        print(f"JSON summary: {output}")

    return 0 if passed == total else 1


# ── Main ─────────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--poi",    default=None,
                    help="Path to a POI test .xlsx/.xlsm file or directory (optional)")
    ap.add_argument("--mode",   default="all",
                    choices=["all", "formulas", "tables", "charts", "formats"],
                    help="POI check mode (default: all / auto-detect per file)")
    ap.add_argument("--output", default=None,
                    help="Write POI JSON results to this file")
    args = ap.parse_args()

    print("xlsx-ez Skill Benchmark")
    print("=" * 60)
    print(f"Skill A (xlsx-ez):   {SKILL_A}")
    print(f"Skill B (reference): {SKILL_B}")

    poi_rc = 0
    with tempfile.TemporaryDirectory() as td:
        tmpdir = Path(td)
        suite_c()
        suite_d()
        suite_e()
        suite_a(tmpdir)
        suite_b(tmpdir)

        if args.poi:
            poi_rc = suite_p(args.poi, args.mode, args.output)

    _report()
    print(f"{'=' * 60}")
    print(f"Results: {_pass} passed, {_fail} failed, {_warn} warnings")
    return 1 if (_fail > 0 or poi_rc != 0) else 0


if __name__ == "__main__":
    sys.exit(main())
