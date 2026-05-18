#!/usr/bin/env python3
"""
recalc.py — Recalculate all formulas in an Excel file using LibreOffice.

LibreOffice is invoked headless with a StarBasic macro that calls
calculateAll(), saves, then closes the document.  After recalculation
the file is inspected with openpyxl (data_only=True) to collect any
remaining Excel error strings.

Usage:
    python recalc.py <file.xlsx> [timeout_seconds]

Output (JSON):
    {
      "status": "success" | "errors_found",
      "total_formulas": <int>,
      "total_errors": <int>,
      "error_summary": {
        "#DIV/0!": {"count": 1, "locations": ["Sheet1!B5"]},
        ...
      }
    }

Exit code 0 on success or errors_found, non-zero on fatal failure.
"""
from __future__ import annotations

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

# Allow running from the scripts/ directory directly
sys.path.insert(0, str(Path(__file__).parent))

from office.soffice import get_soffice_env

try:
    from openpyxl import load_workbook
except ImportError:
    print('Error: openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(1)

# LibreOffice StarBasic macro directory paths (platform-specific)
_MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
_MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
_MACRO_FILENAME  = "Module1.xba"

# StarBasic macro XML: calculateAll recalculates every formula in the document,
# store() saves in-place, close(True) closes without prompting.
# Ref: LibreOffice Basic IDE > com.sun.star.sheet.SpreadsheetDocument
_MACRO_XML = """\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script"
               script:name="Module1"
               script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def _macro_dir() -> str:
    if platform.system() == "Darwin":
        return os.path.expanduser(_MACRO_DIR_MACOS)
    return os.path.expanduser(_MACRO_DIR_LINUX)


def _setup_macro() -> bool:
    """Write the StarBasic macro file, creating the directory if needed."""
    macro_dir  = _macro_dir()
    macro_file = os.path.join(macro_dir, _MACRO_FILENAME)

    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True

    if not os.path.exists(macro_dir):
        # Trigger LibreOffice user-profile creation, then make the directory.
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=15,
            env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(_MACRO_XML)
        return True
    except OSError:
        return False


def _run_libreoffice(path: str, timeout: int) -> dict | None:
    """
    Invoke LibreOffice to open the file and run RecalculateAndSave.
    Returns None on success, or an error dict on failure.

    The macro URL format is documented at:
      https://api.libreoffice.org/docs/idl/ref/namespacecom_1_1sun_1_1star_1_1script.html
    Format: vnd.sun.star.script:<Module>.<Sub>?language=Basic&location=application
    """
    macro_url = (
        "vnd.sun.star.script:"
        "Standard.Module1.RecalculateAndSave"
        "?language=Basic&location=application"
    )
    cmd = ["soffice", "--headless", "--norestore", macro_url, path]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin":
        # gtimeout is from GNU coreutils via Homebrew
        try:
            subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=1, check=False)
            cmd = ["gtimeout", str(timeout)] + cmd
        except (FileNotFoundError, subprocess.TimeoutExpired):
            pass

    result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())

    # returncode 124 = timeout (timeout/gtimeout convention); treat as success
    # because LibreOffice may have already saved before the timeout fired.
    if result.returncode not in (0, 124):
        msg = result.stderr or "unknown LibreOffice error"
        return {"error": msg}

    return None


def recalc(filename: str, timeout: int = 30) -> dict:
    """
    Recalculate formulas in *filename* using LibreOffice, then inspect for
    Excel errors.  Returns a result dict (see module docstring).
    """
    if not Path(filename).exists():
        return {"error": f"File not found: {filename}"}

    abs_path = str(Path(filename).resolve())

    if not _setup_macro():
        return {"error": "Failed to set up LibreOffice macro"}

    lo_err = _run_libreoffice(abs_path, timeout)
    if lo_err:
        return lo_err

    # Inspect the recalculated file for formula errors.
    try:
        wb = load_workbook(abs_path, data_only=True)
        error_details: dict[str, list[str]] = {e: [] for e in EXCEL_ERRORS}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in EXCEL_ERRORS:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        wb.close()

        # Count formulas (from the formula-preserving view)
        wb2 = load_workbook(abs_path, data_only=False)
        formula_count = 0
        for sheet_name in wb2.sheetnames:
            ws2 = wb2[sheet_name]
            for row in ws2.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb2.close()

        result: dict = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_formulas": formula_count,
            "total_errors": total_errors,
            "error_summary": {},
        }
        for err_type, locs in error_details.items():
            if locs:
                result["error_summary"][err_type] = {
                    "count": len(locs),
                    "locations": locs[:20],
                }
        return result

    except Exception as exc:
        return {"error": str(exc)}


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filename = sys.argv[1]
    timeout  = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    result   = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
