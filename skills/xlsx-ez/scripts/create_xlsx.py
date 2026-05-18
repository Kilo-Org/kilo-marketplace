#!/usr/bin/env python3
"""
create_xlsx.py — Create an Excel workbook from a JSON specification.

Usage:
    python create_xlsx.py <spec.json> <output.xlsx>
    python create_xlsx.py --example            # print an example spec

The JSON spec drives workbook construction via openpyxl.  After writing,
run recalc.py to validate formulas with LibreOffice.

Spec format:
{
  "sheets": [
    {
      "name": "Summary",
      "freeze_panes": "A2",
      "column_widths": {"A": 20, "B": 14},
      "rows": [
        {
          "cells": [
            {"value": "Revenue", "bold": true, "bg_color": "4472C4",
             "font_color": "FFFFFF", "number_format": "$#,##0"},
            {"value": "=SUM(Detail!B2:B13)", "number_format": "$#,##0"}
          ]
        }
      ]
    }
  ]
}

Cell fields (all optional):
  value         str | int | float | bool | null
  bold          bool
  italic        bool
  font_size     int (default 10)
  font_name     str (default "Arial")
  font_color    hex RGB string, e.g. "FF0000"
  bg_color      hex RGB solid fill, e.g. "FFFF00"
  number_format Excel format string
  halign        "left" | "center" | "right" | "fill" | "justify"
  valign        "top" | "center" | "bottom"
  wrap_text     bool
  border_bottom "thin" | "medium" | "thick" | "double"
  border_top    "thin" | "medium" | "thick" | "double"
  merge         int  — merge this many columns to the right (e.g. 3 = span 3 cols)
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

EXAMPLE_SPEC = {
    "sheets": [
        {
            "name": "Assumptions",
            "freeze_panes": "A2",
            "column_widths": {"A": 28, "B": 16},
            "rows": [
                {
                    "cells": [
                        {"value": "Parameter",  "bold": True, "bg_color": "4472C4", "font_color": "FFFFFF"},
                        {"value": "Value",       "bold": True, "bg_color": "4472C4", "font_color": "FFFFFF"},
                    ]
                },
                {
                    "cells": [
                        {"value": "Growth Rate", "bg_color": "FFFF00"},
                        {"value": 0.05,          "bg_color": "FFFF00", "number_format": "0.0%"},
                    ]
                },
            ],
        },
        {
            "name": "Forecast",
            "freeze_panes": "B2",
            "column_widths": {"A": 10, "B": 16, "C": 16},
            "rows": [
                {
                    "cells": [
                        {"value": "Year",     "bold": True},
                        {"value": "Revenue",  "bold": True},
                        {"value": "Growth",   "bold": True},
                    ]
                },
                {
                    "cells": [
                        {"value": 2024},
                        {"value": 1000000, "number_format": "$#,##0"},
                        {"value": None},
                    ]
                },
                {
                    "cells": [
                        {"value": 2025},
                        {"value": "=B2*(1+Assumptions!$B$2)", "number_format": "$#,##0"},
                        {"value": "=B3/B2-1", "number_format": "0.0%"},
                    ]
                },
            ],
        },
    ]
}


def _make_font(cell_spec: dict) -> Font:
    return Font(
        name=cell_spec.get("font_name", "Arial"),
        size=cell_spec.get("font_size", 10),
        bold=cell_spec.get("bold", False),
        italic=cell_spec.get("italic", False),
        color=cell_spec.get("font_color", "000000"),
    )


def _make_fill(cell_spec: dict) -> PatternFill | None:
    color = cell_spec.get("bg_color")
    if not color:
        return None
    return PatternFill(fill_type="solid", fgColor=color)


def _make_alignment(cell_spec: dict) -> Alignment | None:
    halign    = cell_spec.get("halign")
    valign    = cell_spec.get("valign")
    wrap_text = cell_spec.get("wrap_text", False)
    if halign or valign or wrap_text:
        return Alignment(
            horizontal=halign,
            vertical=valign,
            wrap_text=wrap_text,
        )
    return None


def _make_border(cell_spec: dict) -> Border | None:
    bottom = cell_spec.get("border_bottom")
    top    = cell_spec.get("border_top")
    if bottom or top:
        return Border(
            bottom=Side(style=bottom) if bottom else Side(),
            top=Side(style=top) if top else Side(),
        )
    return None


def _apply_cell(cell: Any, cell_spec: dict) -> None:
    value = cell_spec.get("value")
    if value is not None:
        cell.value = value

    cell.font = _make_font(cell_spec)

    fill = _make_fill(cell_spec)
    if fill:
        cell.fill = fill

    align = _make_alignment(cell_spec)
    if align:
        cell.alignment = align

    border = _make_border(cell_spec)
    if border:
        cell.border = border

    fmt = cell_spec.get("number_format")
    if fmt:
        cell.number_format = fmt


def build_workbook(spec: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)   # remove default Sheet

    for sheet_spec in spec.get("sheets", []):
        ws = wb.create_sheet(title=sheet_spec.get("name", "Sheet"))

        freeze = sheet_spec.get("freeze_panes")
        if freeze:
            ws.freeze_panes = freeze

        col_widths: dict[str, float] = sheet_spec.get("column_widths", {})
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        pending_merges: list[tuple[int, int, int]] = []

        for row_idx, row_spec in enumerate(sheet_spec.get("rows", []), start=1):
            col_idx = 1
            cells = row_spec.get("cells", [])
            for cell_spec in cells:
                cell = ws.cell(row=row_idx, column=col_idx)
                _apply_cell(cell, cell_spec)

                span = cell_spec.get("merge", 0)
                if span and span > 1:
                    pending_merges.append((row_idx, col_idx, col_idx + span - 1))

                col_idx += 1

            row_height = row_spec.get("height")
            if row_height:
                ws.row_dimensions[row_idx].height = row_height

        for r, c_start, c_end in pending_merges:
            start_coord = f"{get_column_letter(c_start)}{r}"
            end_coord   = f"{get_column_letter(c_end)}{r}"
            ws.merge_cells(f"{start_coord}:{end_coord}")

    return wb


def main() -> None:
    if "--example" in sys.argv:
        print(json.dumps(EXAMPLE_SPEC, indent=2))
        sys.exit(0)

    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    spec_path   = sys.argv[1]
    output_path = sys.argv[2]

    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)

    wb = build_workbook(spec)
    wb.save(output_path)
    print(f"Written: {output_path}")
    print("Run recalc.py to validate formulas with LibreOffice.")


if __name__ == "__main__":
    main()
